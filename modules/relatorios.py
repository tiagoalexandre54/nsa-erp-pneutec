"""
Tela 4 — Relatórios e Exportações.
"""
import streamlit as st
import pandas as pd
import datetime
from io import BytesIO


def tela_relatorios():
    st.title("📈 Relatórios e Exportações")
    st.write("Gere e baixe relatórios em Excel com um clique.")

    df: pd.DataFrame = st.session_state.bd_pneus

    if df.empty:
        st.warning("Nenhuma OS cadastrada. Importe os dados pelo Painel PPCP.")
        return

    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    st.caption(f"Base de dados atualizada em: {agora}  |  Total de OS: {len(df)}")

    st.markdown("---")

    # ── Filtros globais ──────────────────────────────────────────────────────
    with st.expander("🔍 Filtros (aplicados a todos os relatórios)", expanded=False):
        col1, col2 = st.columns(2)
        clientes = ["Todos"] + sorted(df['CLIENTE'].replace('', pd.NA).dropna().unique().tolist())
        filtro_cliente = col1.selectbox("Cliente:", clientes, key="rel_filtro_cliente")

        col2.markdown("**Período (DATA_ENTRADA):**")
        usar_periodo = col2.checkbox("Filtrar por período", key="rel_usar_periodo")
        if usar_periodo:
            c1, c2 = st.columns(2)
            data_ini = c1.date_input("De:", key="rel_data_ini")
            data_fim = c2.date_input("Até:", key="rel_data_fim")

    # Aplica filtros
    df_filtrado = df.copy()
    if filtro_cliente != "Todos":
        df_filtrado = df_filtrado[df_filtrado['CLIENTE'] == filtro_cliente]

    if usar_periodo:
        def _parse_data(s):
            try:
                return pd.to_datetime(s, dayfirst=True, errors='coerce')
            except Exception:
                return pd.NaT

        datas = df_filtrado['DATA_ENTRADA'].apply(_parse_data)
        ini = pd.Timestamp(data_ini)
        fim = pd.Timestamp(data_fim) + pd.Timedelta(days=1)
        df_filtrado = df_filtrado[(datas >= ini) & (datas < fim)]

    st.markdown("---")

    # ── Cards de resumo ──────────────────────────────────────────────────────
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Aguardando",  len(df_filtrado[df_filtrado['STATUS'] == 'Aguardando']))
    col2.metric("Em Limpeza",  len(df_filtrado[df_filtrado['STATUS'] == 'Em Limpeza']))
    col3.metric("Em Produção", len(df_filtrado[df_filtrado['STATUS'] == 'Em Produção']))
    col4.metric("Expedidos",   len(df_filtrado[df_filtrado['STATUS'] == 'Expedido']))
    col5.metric("Total",       len(df_filtrado))

    st.markdown("---")

    # ── Relatório 1: Geral ───────────────────────────────────────────────────
    _card_relatorio(
        titulo="📋 Relatório Geral",
        descricao="Todas as OS com todos os status.",
        df=df_filtrado,
        nome_arquivo="relatorio_geral",
        key="rel_geral"
    )

    # ── Relatório 2: Produzidos ──────────────────────────────────────────────
    df_prod = df_filtrado[df_filtrado['STATUS'] == 'Em Produção'].copy()
    _card_relatorio(
        titulo="🏭 Produzidos (Em Produção)",
        descricao="OS atualmente na linha de produção.",
        df=df_prod,
        nome_arquivo="relatorio_em_producao",
        key="rel_producao"
    )

    # ── Relatório 3: Expedidos (com Lead Time) ───────────────────────────────
    df_exped = df_filtrado[df_filtrado['STATUS'] == 'Expedido'].copy()

    if not df_exped.empty:
        # 1. Converte datas tentando formatos explícitos em ordem (sem dayfirst,
        #    que corromperia datas ISO). Cobre o padrão atual e datas antigas.
        def _to_dt(serie):
            dt = pd.to_datetime(serie, format="%d/%m/%Y %H:%M:%S", errors='coerce')
            for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                falta = dt.isna()
                if not falta.any():
                    break
                dt.loc[falta] = pd.to_datetime(serie[falta], format=fmt, errors='coerce')
            return dt

        entrada_dt = _to_dt(df_exped['DATA_ENTRADA'])
        saida_dt   = _to_dt(df_exped['DATA_SAIDA'])

        # 2. Lead time em dias corridos (1 casa decimal)
        df_exped['LEAD_TIME_DIAS'] = (
            (saida_dt - entrada_dt).dt.total_seconds() / (24 * 3600)
        ).round(1)

        # 3. Média global do período filtrado
        lead_time_medio = df_exped['LEAD_TIME_DIAS'].mean()
    else:
        lead_time_medio = 0.0

    st.subheader("🚚 Expedidos e Performance de Entrega")
    colA, colB = st.columns(2)
    colA.metric("Pneus Expedidos", len(df_exped))
    colB.metric(
        "Lead Time Médio (Dias)",
        f"{lead_time_medio:.1f}" if not pd.isna(lead_time_medio) else "0.0",
    )

    _card_relatorio(
        titulo="Detalhes da Expedição",
        descricao="OS que já saíram para entrega, incluindo o tempo de ciclo (Lead Time).",
        df=df_exped,
        nome_arquivo="relatorio_expedidos_leadtime",
        key="rel_expedidos"
    )

    # ── Relatório 4: Romaneio por Cliente ────────────────────────────────────
    st.subheader("📦 Romaneio de Carga por Cliente")
    st.write("Selecione o cliente para gerar o romaneio com os pneus expedidos.")

    clientes_rom = sorted(
        df_filtrado[df_filtrado['STATUS'] == 'Expedido']['CLIENTE']
        .replace('', pd.NA).dropna().unique().tolist()
    )

    if not clientes_rom:
        st.info("Nenhum pneu expedido no período/filtro selecionado.")
    else:
        cliente_rom = st.selectbox("Cliente para o Romaneio:", clientes_rom, key="rel_rom_cliente")
        df_rom = df_filtrado[
            (df_filtrado['CLIENTE'] == cliente_rom) &
            (df_filtrado['STATUS'] == 'Expedido')
        ].copy()

        st.dataframe(
            df_rom[['NRORDEM', 'NRSERIE', 'DESENHO', 'DATA_ENTRADA', 'DATA_SAIDA']],
            use_container_width=True, hide_index=True
        )
        st.caption(f"Total: {len(df_rom)} pneu(s)")

        _botoes_download(
            df=df_rom,
            nome_arquivo=f"romaneio_{cliente_rom.replace(' ', '_')}",
            key="rel_romaneio"
        )

    st.markdown("---")

    # ── Exportação completa (todos os relatórios em abas) ────────────────────
    st.subheader("📁 Exportação Completa (Excel com todas as abas)")
    st.write("Gera um único arquivo Excel com abas separadas para cada relatório.")

    if st.button("📥 Gerar Excel Completo", key="btn_excel_completo", type="primary"):
        xlsx = _gerar_excel_completo(df_filtrado)
        nome = f"NSA_Relatorio_Completo_{datetime.datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        st.download_button(
            label="⬇️ Baixar Excel Completo",
            data=xlsx,
            file_name=nome,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_completo"
        )


# ── Helpers ──────────────────────────────────────────────────────────────────

def _card_relatorio(titulo: str, descricao: str, df: pd.DataFrame, nome_arquivo: str, key: str):
    st.subheader(titulo)
    st.caption(descricao)

    if df.empty:
        st.info("Nenhuma OS nesta categoria com os filtros aplicados.")
        st.markdown("---")
        return

    st.dataframe(df, use_container_width=True, hide_index=True)
    st.caption(f"Total: {len(df)} OS")
    _botoes_download(df, nome_arquivo, key)
    st.markdown("---")


def _botoes_download(df: pd.DataFrame, nome_arquivo: str, key: str):
    """Exibe botões de download em CSV e Excel lado a lado."""
    col1, col2 = st.columns([1, 1])

    # CSV
    csv_bytes = df.to_csv(index=False).encode('utf-8-sig')  # utf-8-sig = abre certo no Excel BR
    col1.download_button(
        label="⬇️ Baixar CSV",
        data=csv_bytes,
        file_name=f"{nome_arquivo}_{datetime.datetime.now().strftime('%d%m%Y')}.csv",
        mime="text/csv",
        key=f"{key}_csv"
    )

    # Excel
    xlsx_bytes = _df_para_excel(df, nome_arquivo)
    col2.download_button(
        label="⬇️ Baixar Excel",
        data=xlsx_bytes,
        file_name=f"{nome_arquivo}_{datetime.datetime.now().strftime('%d%m%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"{key}_xlsx"
    )


def _df_para_excel(df: pd.DataFrame, nome_aba: str = "Dados") -> bytes:
    """Converte DataFrame para bytes Excel com formatação básica."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nome_aba[:31])
        _formatar_planilha(writer.sheets[nome_aba[:31]], df)
    return buf.getvalue()


def _gerar_excel_completo(df: pd.DataFrame) -> bytes:
    """Gera Excel com múltiplas abas."""
    buf = BytesIO()
    abas = {
        "Geral":        df,
        "Em Limpeza":   df[df['STATUS'] == 'Em Limpeza'],
        "Em Producao":  df[df['STATUS'] == 'Em Produção'],
        "Expedidos":    df[df['STATUS'] == 'Expedido'],
        "Aguardando":   df[df['STATUS'] == 'Aguardando'],
    }

    # Aba de Romaneio por cliente
    clientes_exped = df[df['STATUS'] == 'Expedido']['CLIENTE'].replace('', pd.NA).dropna().unique()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nome_aba, df_aba in abas.items():
            df_aba.to_excel(writer, index=False, sheet_name=nome_aba)
            _formatar_planilha(writer.sheets[nome_aba], df_aba)

        for cliente in sorted(clientes_exped):
            df_cli = df[(df['CLIENTE'] == cliente) & (df['STATUS'] == 'Expedido')]
            nome_aba = f"ROM_{cliente[:24]}"
            df_cli.to_excel(writer, index=False, sheet_name=nome_aba)
            _formatar_planilha(writer.sheets[nome_aba], df_cli)

    return buf.getvalue()


def _formatar_planilha(ws, df: pd.DataFrame):
    """Ajusta largura das colunas e destaca o cabeçalho."""
    from openpyxl.styles import PatternFill, Font, Alignment

    header_fill = PatternFill("solid", fgColor="003366")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

        # Largura automática baseada no conteúdo
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if not df.empty else 0
        )
        ws.column_dimensions[cell.column_letter].width = min(max_len + 4, 50)

    # Congela a linha de cabeçalho
    ws.freeze_panes = "A2"
