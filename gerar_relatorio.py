from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

wb = Workbook()

azul_esc = PatternFill('solid', fgColor='003366')
azul_med = PatternFill('solid', fgColor='1a5276')
azul_cla = PatternFill('solid', fgColor='d6eaf8')
branco   = PatternFill('solid', fgColor='ffffff')
fonte_br = Font(bold=True, color='ffffff', size=11)
fonte_az = Font(bold=True, color='003366', size=11)
fonte_n  = Font(color='000000', size=10)
centro   = Alignment(horizontal='center', vertical='center', wrap_text=True)
esq      = Alignment(horizontal='left', vertical='center', wrap_text=True)

def titulo(ws, texto, ncols=2):
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    c = ws['A1']
    c.value = texto
    c.fill = azul_esc
    c.font = Font(bold=True, color='ffffff', size=13)
    c.alignment = centro
    ws.row_dimensions[1].height = 30

def header_row(ws, headers, row=2):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = azul_med
        c.font = fonte_br
        c.alignment = centro

def data_row(ws, row, values):
    fill = azul_cla if row % 2 == 0 else branco
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = fonte_n
        c.alignment = esq
    ws.row_dimensions[row].height = 22

def secao(ws, row, texto, ncols=2):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = azul_med
    c.font = fonte_br
    c.alignment = esq


# ══════════════════════════════════
# ABA 1 — VISÃO GERAL
# ══════════════════════════════════
ws1 = wb.active
ws1.title = 'Visão Geral'
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 65
titulo(ws1, 'NSA PNEUTEC — RELATÓRIO TÉCNICO DO SISTEMA ERP', 2)

agora = datetime.now().strftime("%d/%m/%Y %H:%M")
ws1.merge_cells('A2:B2')
c = ws1['A2']
c.value = 'Gerado em: ' + agora
c.fill = azul_med
c.font = fonte_br
c.alignment = centro

dados = [
    ('INFORMAÇÕES DO PROJETO', ''),
    ('Nome do Sistema', 'ERP Reformadora de Pneus — MES / PPCP'),
    ('Empresa', 'NSA PNEUTEC — Desde 1951'),
    ('Finalidade', 'Gestão de produção, recebimento, expedição e rastreamento de pneus recauchutados'),
    ('Data de Início', '01/06/2026'),
    ('Data do Relatório', datetime.now().strftime('%d/%m/%Y')),
    ('Total de Telas', '6 módulos: PPCP, Recebimento, Entrada, Expedição, Relatórios, Acesso Mobile'),
    ('Total de Linhas Python', '2.097 linhas de código'),
    ('Tamanho do Projeto', '1.080 KB'),
    ('', ''),
    ('HOSPEDAGEM', ''),
    ('Ambiente Local', 'Windows — http://192.168.15.9:3001'),
    ('Ambiente Nuvem', 'https://nsa-erp-pneutec-jwmesdmiv9peeipz56njt7.streamlit.app'),
    ('Repositório GitHub', 'https://github.com/tiagoalexandre54/nsa-erp-pneutec'),
    ('Início Automático', 'Windows Startup — inicia com o computador automaticamente'),
]

row = 3
for k, v in dados:
    if v == '':
        secao(ws1, row, k, 2)
    else:
        c1 = ws1.cell(row=row, column=1, value=k)
        c2 = ws1.cell(row=row, column=2, value=v)
        fill = azul_cla if row % 2 == 0 else branco
        c1.fill = fill; c1.font = fonte_az; c1.alignment = esq
        c2.fill = fill; c2.font = fonte_n;  c2.alignment = esq
    ws1.row_dimensions[row].height = 22
    row += 1


# ══════════════════════════════════
# ABA 2 — TECNOLOGIAS
# ══════════════════════════════════
ws2 = wb.create_sheet('Tecnologias')
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 50
titulo(ws2, 'TECNOLOGIAS E BIBLIOTECAS UTILIZADAS', 4)
header_row(ws2, ['Tecnologia', 'Versão', 'Categoria', 'Finalidade'])

tecnologias = [
    ('Python',           '3.12.10',  'Linguagem',    'Linguagem principal de todo o sistema'),
    ('Streamlit',        '1.58.0',   'Framework Web','Interface: telas, componentes, navegação, sidebar'),
    ('Pandas',           '3.0.3',    'Dados',        'Manipulação e processamento do banco de dados CSV'),
    ('pdfplumber',       '0.11.9',   'PDF',          'Leitura e extração de dados dos PDFs de pedido de recapagem'),
    ('openpyxl',         '3.1.5',    'Excel',        'Geração de relatórios Excel formatados com múltiplas abas'),
    ('qrcode + Pillow',  '8.2',      'Mobile',       'Geração de QR Codes para acesso pelo celular'),
    ('requests',         '2.34.2',   'Nuvem',        'Comunicação com API do GitHub para salvar dados na nuvem'),
    ('pyngrok',          '8.1.2',    'Rede',         'Suporte a túnel SSH para link público externo'),
    ('Git',              '2.54',     'Versionamento','Controle de versão e deploy automático no Streamlit Cloud'),
    ('GitHub API',       'v3',       'Nuvem',        'Leitura e escrita do CSV de dados no repositório remoto'),
    ('Streamlit Cloud',  'Free',     'Hospedagem',   'Servidor gratuito para acesso 24h de qualquer lugar'),
    ('Windows Startup',  '-',        'SO Windows',   'Inicialização automática do app ao ligar o computador'),
    ('SSH / Serveo.net', '-',        'Rede',         'Túnel público sem conta para acesso externo temporário'),
    ('VBS Script',       '-',        'Windows',      'Execução silenciosa do app em background no Windows'),
    ('Regex (re)',       'builtin',  'Python',       'Parser robusto para extração de dados do PDF'),
    ('pathlib / os',    'builtin',  'Python',       'Manipulação de caminhos de arquivos multiplataforma'),
    ('tempfile',        'builtin',  'Python',       'Arquivos temporários para impressão e importação'),
    ('base64',          'builtin',  'Python',       'Encoding para envio de arquivos via GitHub API'),
    ('subprocess',      'builtin',  'Python',       'Execução do processo SSH para túnel público'),
    ('socket',          'builtin',  'Python',       'Detecção do IP local da rede Wi-Fi'),
]

for i, row_data in enumerate(tecnologias, start=3):
    data_row(ws2, i, row_data)


# ══════════════════════════════════
# ABA 3 — MÓDULOS
# ══════════════════════════════════
ws3 = wb.create_sheet('Módulos do Sistema')
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 24
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 55
titulo(ws3, 'MÓDULOS E FUNCIONALIDADES DO SISTEMA', 5)
header_row(ws3, ['Módulo', 'Arquivo', 'Linhas', 'Tela', 'Funcionalidades Principais'])

modulos = [
    ('Ponto de Entrada',   'app.py',                 58,  'Todas',        'Roteamento entre telas, sidebar com logo, mini-painel de status, QR Code mobile'),
    ('Banco de Dados',     'modules/database.py',   303,  'Interno',      '40+ mapeamentos de colunas, importação CSV/nuvem, escrita atômica, normalização de status ERP'),
    ('Painel PPCP',        'modules/painel_pcp.py', 275,  'PPCP',         'Indicadores tempo real, tabela colorida, filtros, importar CSV/PDF, excluir OS, zerar sistema'),
    ('Recebimento',        'modules/recebimento.py',697,  'Recebimento',  'Mapa galpão 8x5 pallets, bipe QR, rastreamento OS, formulário manual, impressão HTML/PDF'),
    ('Entrada Produção',   'modules/entrada.py',    127,  'Entrada',      'Alerta coletas pendentes por cliente, bipe entrada produção, tabela de OS em produção'),
    ('Expedição',          'modules/expedicao.py',  148,  'Expedição',    'Painel de embarque por cliente, bipe com trava anti-troca, romaneio de carga'),
    ('Relatórios',         'modules/relatorios.py', 194,  'Relatórios',   'Geral/Produzidos/Expedidos/Romaneio, filtro período e cliente, CSV e Excel formatado'),
    ('Importação PDF',     'modules/pdf_import.py', 138,  'Interno',      'Parser por medida do pneu (âncora), séries alfanuméricas, cruzamento automático com CSV'),
    ('Acesso Mobile',      'modules/acesso_mobile.py',157,'Sidebar',      'QR Code Wi-Fi local, túnel Serveo público, detecção automática local/nuvem'),
]

for i, row_data in enumerate(modulos, start=3):
    data_row(ws3, i, row_data)
    ws3.row_dimensions[i].height = 35


# ══════════════════════════════════
# ABA 4 — HISTÓRICO
# ══════════════════════════════════
ws4 = wb.create_sheet('Histórico de Versões')
ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 75
titulo(ws4, 'HISTÓRICO DE ATUALIZAÇÕES — DO INÍCIO AO MOMENTO ATUAL', 3)
header_row(ws4, ['Data', 'Versão', 'Descrição da Atualização'])

historico = [
    ('01/06/2026', 'v1.0', 'Criação da estrutura base: Painel PPCP, Entrada, Expedição — banco CSV local, bipe QR, trava anti-troca'),
    ('01/06/2026', 'v1.1', 'Importação CSV externo com 40+ mapeamentos de colunas e datas de diferentes sistemas ERP'),
    ('01/06/2026', 'v1.2', 'Importação de PDF (Pedido de Recapagem): extração de cliente, datas, modelo e pneus'),
    ('01/06/2026', 'v1.3', 'Correção parser PDF para séries alfanuméricas (ex: 1223A) e desenhos sem traço (VL110LA)'),
    ('01/06/2026', 'v1.4', 'NRORDEM mapeado para IDITEMPEDIDOPNEU (ID único por pneu); IDPEDIDOPNEU como coluna própria'),
    ('01/06/2026', 'v1.5', 'Alerta visual de coletas pendentes na tela de Entrada; cruzamento automático PDF × CSV por NRSERIE'),
    ('01/06/2026', 'v1.6', 'Painel de pneus prontos para embarque na Expedição com filtro por cliente'),
    ('01/06/2026', 'v1.7', 'Módulo de Relatórios: Geral, Em Produção, Expedidos, Romaneio — CSV e Excel multi-abas formatado'),
    ('01/06/2026', 'v1.8', 'Acesso Mobile: QR Code automático Wi-Fi local + túnel público via Serveo sem conta'),
    ('01/06/2026', 'v1.9', 'Instalador Windows (INSTALAR.bat) + início automático com o SO + atalho Área de Trabalho'),
    ('01/06/2026', 'v1.10','Pacote de distribuição ZIP (GERAR_PACOTE.bat) para instalação em outros computadores'),
    ('02/06/2026', 'v2.0', 'Módulo Recebimento: mapa visual 8 fileiras × 5 pallets, capacidade 8 pneus cada'),
    ('02/06/2026', 'v2.1', 'Recebimento: bipe QR Code para alocar OS ao pallet, rastreamento completo por NRORDEM'),
    ('02/06/2026', 'v2.2', 'Recebimento: formulário manual de inserção de coleta com validação de capacidade'),
    ('02/06/2026', 'v2.3', 'Recebimento: impressão HTML do mapa do galpão, pallet individual e todos os pallets'),
    ('02/06/2026', 'v2.4', 'Correção TypeError NaN em LOCAL_PALLET; escrita atômica do CSV; deduplicação de colunas'),
    ('02/06/2026', 'v2.5', 'Hospedagem na nuvem: GitHub + Streamlit Cloud; banco de dados via GitHub API REST'),
    ('02/06/2026', 'v2.6', 'Correção compatibilidade Linux/Streamlit Cloud; _salvar() adaptado para local e nuvem'),
    ('02/06/2026', 'v2.7', 'URL pública definida: nsa-erp-pneutec-jwmesdmiv9peeipz56njt7.streamlit.app'),
    ('03/06/2026', 'v2.8', 'Sidebar renomeada: "PPCP — Planejamento, Programação e Controle da Produção"'),
    ('03/06/2026', 'v2.9', 'Correção TypeError na ordenação de pallets com formatos inesperados de chave'),
    ('03/06/2026', 'v3.0', 'Geração deste Relatório Técnico completo em Excel'),
]

for i, row_data in enumerate(historico, start=3):
    data_row(ws4, i, row_data)


# ══════════════════════════════════
# ABA 5 — INFRAESTRUTURA
# ══════════════════════════════════
ws5 = wb.create_sheet('Infraestrutura')
ws5.column_dimensions['A'].width = 32
ws5.column_dimensions['B'].width = 60
titulo(ws5, 'INFRAESTRUTURA E ARQUITETURA DO SISTEMA', 2)

infra = [
    ('ARQUITETURA', ''),
    ('Tipo de Aplicação',  'Web Local + Cloud (Streamlit Framework)'),
    ('Padrão Arquitetural','Módulos separados por responsabilidade (similar MVC)'),
    ('Banco de Dados',     'CSV local (data/ordens.csv) + GitHub API na nuvem'),
    ('Comunicação',        'HTTP/WebSocket interno (Streamlit) + HTTPS (GitHub API)'),
    ('', ''),
    ('AMBIENTE LOCAL', ''),
    ('Sistema Operacional','Windows 10 / 11'),
    ('Python',             '3.12.10'),
    ('Porta',              '3001'),
    ('Endereço de Escuta', '0.0.0.0 (aceita conexões de toda a rede local)'),
    ('IP de Acesso Rede',  'http://192.168.15.9:3001'),
    ('Inicialização',      'Automática via Windows Startup — iniciar_silencioso.vbs'),
    ('Instalação Outros PCs','INSTALAR.bat — baixa Python, dependências e configura tudo'),
    ('', ''),
    ('AMBIENTE NUVEM', ''),
    ('Plataforma',         'Streamlit Community Cloud (gratuito)'),
    ('Repositório',        'GitHub — tiagoalexandre54/nsa-erp-pneutec (público)'),
    ('Branch Principal',   'main'),
    ('Deploy',             'Automático a cada git push (CI/CD nativo)'),
    ('URL Pública',        'https://nsa-erp-pneutec-jwmesdmiv9peeipz56njt7.streamlit.app'),
    ('Banco na Nuvem',     'GitHub API REST — lê/escreve data/ordens.csv no repositório'),
    ('Secrets (senhas)',   'Armazenados com criptografia no Streamlit Cloud Secrets'),
    ('', ''),
    ('SEGURANÇA', ''),
    ('Token GitHub',       'PAT com escopo repo — armazenado em secrets (não exposto no código)'),
    ('Dados',              'Repositório público (código) — dados sensíveis só nos secrets'),
    ('Trava Expedição',    'Validação cliente × OS antes de qualquer expedição'),
    ('Confirmação Exclusão','Duplo checkbox obrigatório para excluir ou zerar o sistema'),
    ('', ''),
    ('ESTRUTURA DE ARQUIVOS', ''),
    ('app.py',                     'Ponto de entrada — roteamento e sidebar'),
    ('modules/database.py',        'Camada de dados — CSV local e GitHub API'),
    ('modules/painel_pcp.py',      'Tela PPCP — importações e indicadores'),
    ('modules/recebimento.py',     'Tela Recebimento — mapa de pallets completo'),
    ('modules/entrada.py',         'Tela Entrada — bipe de produção'),
    ('modules/expedicao.py',       'Tela Expedição — controle de carga'),
    ('modules/relatorios.py',      'Tela Relatórios — exportações'),
    ('modules/pdf_import.py',      'Parser de PDF — extração de pedidos de recapagem'),
    ('modules/acesso_mobile.py',   'QR Code e link público'),
    ('data/ordens.csv',            'Banco de dados principal'),
    ('assets/logo.png',            'Logo NSA Pneutec 75 Anos'),
    ('.streamlit/config.toml',     'Configurações tema e comportamento'),
    ('.streamlit/secrets.toml',    'Tokens e configurações privadas (não vai ao GitHub)'),
    ('INSTALAR.bat',               'Instalador automático para novos PCs'),
    ('RODAR.bat',                  'Inicializador manual do sistema'),
    ('GERAR_PACOTE.bat',           'Gera ZIP para distribuição em outros PCs'),
    ('iniciar_silencioso.vbs',     'Script de inicialização silenciosa no Windows'),
]

row = 3
for k, v in infra:
    if v == '':
        secao(ws5, row, k, 2)
    else:
        fill = azul_cla if row % 2 == 0 else branco
        c1 = ws5.cell(row=row, column=1, value=k)
        c2 = ws5.cell(row=row, column=2, value=v)
        c1.fill = fill; c1.font = fonte_az; c1.alignment = esq
        c2.fill = fill; c2.font = fonte_n;  c2.alignment = esq
    ws5.row_dimensions[row].height = 20
    row += 1

# ── Salva na Área de Trabalho ──
desktop = os.path.join(os.path.expanduser('~'), 'OneDrive', 'Desktop')
if not os.path.exists(desktop):
    desktop = os.path.expanduser('~')
path = os.path.join(desktop, 'Relatorio_Tecnico_ERP_NSA_Pneutec.xlsx')
wb.save(path)
print('Salvo em: ' + path)
